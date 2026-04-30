"""XLSX -> markdown via openpyxl.

Each sheet becomes an `## Sheet: name` heading followed by a markdown table.
For sheets with >500 rows, only the first 500 are emitted with a "(... N more rows)" footer.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROW_LIMIT = 500


def _cell(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("|", "\\|")
    return s.strip()


def convert(path: Path) -> tuple[str, str]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"## Sheet: {sheet_name}")
        # Find first non-empty row to use as header
        header_idx = 0
        for i, r in enumerate(rows):
            if r and any(c is not None and str(c).strip() for c in r):
                header_idx = i
                break
        header = rows[header_idx]
        ncols = len(header)
        if ncols == 0:
            continue
        parts.append("| " + " | ".join(_cell(c) for c in header) + " |")
        parts.append("| " + " | ".join("---" for _ in range(ncols)) + " |")
        data_rows = rows[header_idx + 1 :]
        emitted = 0
        for r in data_rows:
            if emitted >= ROW_LIMIT:
                parts.append(f"\n*(... {len(data_rows) - emitted} more rows truncated)*\n")
                break
            cells = [_cell(c) for c in (list(r) + [None] * (ncols - len(r)))[:ncols]]
            if not any(cells):
                continue
            parts.append("| " + " | ".join(cells) + " |")
            emitted += 1
    if not parts:
        raise ValueError("XLSX had no sheets with data")
    return "\n\n".join(parts), "openpyxl"
