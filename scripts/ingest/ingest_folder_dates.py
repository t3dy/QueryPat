"""
Improve segment and document dating using Exegesis organization.xlsx.

Reads: Exegesis organization.xlsx (Folder pagecounts sheet)
Updates: segments and documents tables with approximate dates from folder research.

This script:
1. Builds a folder-number → date-range mapping from the spreadsheet
2. Maps section numbers to Exegesis folders using known correspondences
3. Assigns approximate dates to undated segments based on their section/folder
4. Fixes orphaned segments (DOC_EXEG_None) by reassigning to correct parent docs
5. Updates section_order on documents for chronological ordering
"""

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from date_norms import normalize_date


def load_folder_dates(xlsx_path: Path) -> dict:
    """Load folder → date mapping from the spreadsheet."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb['Folder pagecounts and compariso']

    folder_dates = {}
    for r in range(2, ws.max_row + 1):
        folder_num = ws.cell(r, 1).value
        if folder_num is None:
            continue

        approx = ws.cell(r, 2).value
        pj_jl = ws.cell(r, 5).value
        earliest = ws.cell(r, 6).value
        latest = ws.cell(r, 7).value
        notes = ws.cell(r, 3).value

        # Parse earliest/latest dates
        def parse_date(d):
            if d is None:
                return None
            if isinstance(d, datetime):
                return d.strftime('%Y-%m-%d')
            s = str(d).strip()
            if s in ('??', '', 'None'):
                return None
            return s

        e = parse_date(earliest)
        l = parse_date(latest)

        # Build display string
        if e and l:
            # Extract years
            ey = e[:4] if e else None
            ly = l[:4] if l else None
            em = e[5:7] if len(e) >= 7 else None
            lm = l[5:7] if len(l) >= 7 else None

            if ey == ly and em == lm:
                # Same month
                nd = normalize_date(e)
                display = nd.date_display
                confidence = 'approximate'
            elif ey == ly:
                # Same year
                display = f"circa {ey}"
                confidence = 'circa'
            else:
                display = f"circa {ey}-{ly}"
                confidence = 'circa'
        elif e:
            nd = normalize_date(e)
            display = f"circa {nd.date_display}"
            confidence = 'circa'
        else:
            display = None
            confidence = None

        folder_dates[int(folder_num)] = {
            'earliest': e,
            'latest': l,
            'display': display,
            'confidence': confidence,
            'approx_text': str(approx) if approx else None,
            'pj_jl': str(pj_jl) if pj_jl else None,
            'notes': str(notes) if notes else None,
        }

    wb.close()
    return folder_dates


# Known mapping: our section numbers → Exegesis folder numbers
# Based on the content (letters to specific people at known dates, matching folder 5 = letters 2/75-9/76)
# Section 1-13 are from Folder 5 (letters 2/75-9/76)
# Section 14-15 are from folders 23-36 area (late 1976-1977)
# Section 16 is the massive 1978 journal section (folders 2,3,9,11,14-22,28-30,38,50,52,58)
# Section 17 is 1981 (folders 49,56,59-91 area)

# For undated letters (sections 3,4,8,9,10) - these are all in Folder 5
# which spans 1975-02-27 to 1976-09-30
SECTION_TO_FOLDER = {
    1: 5,    # Claudia letter, Feb 27 1975
    2: 5,    # Phyllis letter, Mar 2 1975
    3: 5,    # Tony letter (undated) - folder 5 spans Feb 1975-Sep 1976
    4: 5,    # Claudia letter (undated, Folder 4 is 1974-1975 but section 4 is a letter)
    5: 5,    # Ursula letter, Mar 5 1975
    6: 5,    # Journal entry, Mar 4 1975
    7: 5,    # Henry letter, Feb 29 1975
    8: 5,    # Claudia letter (undated)
    9: 5,    # Claudia letter (undated)
    10: 5,   # Claudia letter (undated) - long undated section
    11: 5,   # Doris letter, Sep 17 1975
    12: 5,   # Journal, Sep 15 1975
    13: 5,   # Journal, Nov 5 1975
    14: 23,  # Dorothy letter, Sep 12 1976 -> folder 23 = Nov-Dec 1976
    15: None, # Dorothy letter massive - spans multiple folders 1976-1977
    16: None, # Massive 1978 journal - spans many folders
    17: None, # Pat letter 1981 - spans many folders
}


def run(db: sqlite3.Connection, source_dir: Path):
    print("Improving dates from Exegesis organization.xlsx...")

    xlsx_path = Path('C:/QueryPat/Exegesis organization.xlsx')
    if not xlsx_path.exists():
        print(f"  SKIP: {xlsx_path} not found")
        return

    folder_dates = load_folder_dates(xlsx_path)
    print(f"  Loaded date info for {len(folder_dates)} folders")

    updated_segs = 0
    updated_docs = 0
    reassigned = 0

    # 1. Fix orphaned segments in DOC_EXEG_None that have dates in their names
    orphans = db.execute("""
        SELECT seg_id, title FROM segments WHERE doc_id = 'DOC_EXEG_None'
    """).fetchall()

    for seg_id, title in orphans:
        # These have titles like "1978-10-10_SECTION_016_136.txt"
        if title and 'SECTION_' in title:
            # Extract section number
            import re
            m = re.search(r'SECTION_(\d+)', title)
            if m:
                section_num = int(m.group(1))
                doc_id = f'DOC_EXEG_SECTION_{section_num:03d}'
                # Check if parent exists
                exists = db.execute("SELECT 1 FROM documents WHERE doc_id = ?", (doc_id,)).fetchone()
                if exists:
                    db.execute("UPDATE segments SET doc_id = ? WHERE seg_id = ?", (doc_id, seg_id))
                    reassigned += 1

    if reassigned:
        db.commit()
        print(f"  Reassigned {reassigned} orphaned segments to correct parent documents")

    # 2. Date undated sections based on folder mapping
    for section_num, folder_num in SECTION_TO_FOLDER.items():
        if folder_num is None:
            continue

        doc_id = f'DOC_EXEG_SECTION_{section_num:03d}'
        doc = db.execute("SELECT date_start FROM documents WHERE doc_id = ?", (doc_id,)).fetchone()
        if not doc:
            continue

        # Only update if document has no date
        if doc[0] is not None:
            continue

        fd = folder_dates.get(folder_num)
        if not fd or not fd['earliest']:
            continue

        db.execute("""
            UPDATE documents SET
                date_start = ?,
                date_end = ?,
                date_display = ?,
                date_confidence = ?,
                date_basis = ?
            WHERE doc_id = ? AND date_start IS NULL
        """, (
            fd['earliest'], fd['latest'],
            fd['display'], fd['confidence'],
            f"Exegesis folder {folder_num} date range",
            doc_id,
        ))
        updated_docs += 1

    # 3. Date undated segments in sections that now have folder-based dates
    undated_segs = db.execute("""
        SELECT s.seg_id, s.doc_id, d.date_start, d.date_end, d.date_display, d.date_confidence
        FROM segments s
        JOIN documents d ON s.doc_id = d.doc_id
        WHERE s.date_start IS NULL AND d.date_start IS NOT NULL
    """).fetchall()

    for seg_id, doc_id, d_start, d_end, d_display, d_conf in undated_segs:
        db.execute("""
            UPDATE segments SET
                date_start = ?,
                date_end = ?,
                date_display = ?,
                date_confidence = 'inferred',
                date_basis = 'Inherited from parent document folder date'
            WHERE seg_id = ?
        """, (d_start, d_end, d_display, seg_id))
        updated_segs += 1

    # 4. Also fix segments that have dates in their ID but null date fields
    # (e.g., DOC_EXEG_None segments with "1978-10-10" in their name)
    import re
    null_dated = db.execute("""
        SELECT seg_id, title FROM segments WHERE date_start IS NULL
    """).fetchall()

    for seg_id, title in null_dated:
        if not title:
            continue
        m = re.match(r'(\d{4}-\d{2}-\d{2})', title)
        if m:
            date_str = m.group(1)
            nd = normalize_date(date_str)
            db.execute("""
                UPDATE segments SET
                    date_start = ?,
                    date_display = ?,
                    date_confidence = 'approximate',
                    date_basis = 'Extracted from segment filename'
                WHERE seg_id = ?
            """, (nd.date_start, nd.date_display, seg_id))
            updated_segs += 1

    db.commit()

    # Report
    remaining = db.execute("SELECT COUNT(*) FROM segments WHERE date_start IS NULL").fetchone()[0]
    print(f"  Updated {updated_docs} document dates, {updated_segs} segment dates")
    print(f"  Remaining undated segments: {remaining}")


if __name__ == '__main__':
    db_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('C:/QueryPat/database/unified.sqlite')
    db = sqlite3.connect(str(db_path))
    run(db, Path('C:/ExegesisAnalysis'))
    db.close()
